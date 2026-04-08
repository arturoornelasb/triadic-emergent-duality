"""Update Excel with V8_xray zoom2 results — cascade confirmed cross-architecture."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
from datetime import date
import openpyxl
from openpyxl.styles import PatternFill

EXCEL_PATH = r'C:\Github\7_caminos\seven-bridges\analisis_cruzado\analisis_completo_danza_cosmica.xlsx'
GREEN = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
today = date.today().isoformat()

wb = openpyxl.load_workbook(EXCEL_PATH)

# --- 1. Update Roadmap ---
ws = wb['Roadmap']
status_col = None
for col in range(1, ws.max_column + 1):
    for check_row in [1, 2, 3]:
        val = ws.cell(row=check_row, column=col).value
        if val and 'Status' in str(val):
            status_col = col
            break
    if status_col:
        break

if status_col:
    # Update #43 to COMPLETED
    for row in range(3, ws.max_row + 1):
        num = ws.cell(row=row, column=1).value
        if num == 43:
            ws.cell(row=row, column=status_col,
                value=f'COMPLETADO {today}. V8_xray zoom2 (5-step res): '
                      f'CASCADA CONFIRMADA en v8 (GPT-2 Medium). '
                      f'65/72 bits cambian en oleadas de 1-7 bits en ~145 steps. '
                      f'7 bits scaffold estables. Churn max 0.17, nunca 1.0. '
                      f'Patron identico a v9 pero cascada mas larga (modelo mas grande).')
            ws.cell(row=row, column=status_col).fill = GREEN
            break

    # Add #44: cross-architecture cascade confirmation
    exists = False
    for row in range(3, ws.max_row + 1):
        num = ws.cell(row=row, column=1).value
        if num == 44:
            exists = True
            target_row = row
            break
    if not exists:
        target_row = ws.max_row + 1
        ws.cell(row=target_row, column=1, value=44)
        ws.cell(row=target_row, column=2,
            value='Confirmacion cross-arquitectura del patron de cascada')
    ws.cell(row=target_row, column=status_col,
        value=f'COMPLETADO {today}. Cascada confirmada en 2 modelos: '
              f'v9 (GPT-Neo 125M, 31/72 bits, ~45 steps) y '
              f'v8 (GPT-2 Medium, 65/72 bits, ~145 steps). '
              f'Churn real ~0.17 en ambos. La "explosion instantanea" era artefacto '
              f'de baja resolucion. El patron de cascada es universal.')
    ws.cell(row=target_row, column=status_col).fill = GREEN
    print(f'Roadmap #43 updated, #44 {"updated" if exists else "added"}')

# --- 2. Update Status formalizacion Q5 ---
ws = wb['Status formalizacion']
probes_start = None
for row in range(1, ws.max_row + 1):
    val = ws.cell(row=row, column=1).value
    if val and 'SECCION D' in str(val):
        probes_start = row
        break

if probes_start:
    header_row = probes_start + 1
    # Find v8 column (or v9 column to add cross-arch note)
    v9_col = None
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val and 'v9' in str(val).lower():
            v9_col = col
            break

    if v9_col:
        for row in range(header_row + 1, ws.max_row + 1):
            val = ws.cell(row=row, column=1).value
            if val and 'Q5' in str(val):
                old = str(ws.cell(row=row, column=v9_col).value or '')
                if 'cross-arch' not in old.lower():
                    ws.cell(row=row, column=v9_col,
                        value=old + f' | CROSS-ARCH ({today}): Cascada confirmada en v8 '
                              f'(GPT-2 Medium). 65/72 bits, ~145 steps. '
                              f'Patron universal, no artefacto de un modelo.')
                break
        print('Q5 updated with cross-arch confirmation')

# --- 3. Documentacion ---
ws = wb['Documentacion'] if 'Documentacion' in wb.sheetnames else None
if ws:
    entries = [
        ('train_v8_xray_zoom2.bat', 'model/',
         f'Paso 3 V8_xray: 30K->30150 cada 5 steps. Ultra-zoom en la explosion.'),
        ('save_v8_xray_zoom2_timeline.py', 'model/',
         f'Extraccion reptimeline para V8_xray paso 3. Incluye mapeo cascada-a-capas.'),
        ('update_excel_v8_xray.py', 'model/',
         f'Excel update: V8_xray paso 1 results.'),
        ('update_excel_v8_xray_zoom2.py', 'model/',
         f'Excel update: V8_xray zoom2 cascade confirmation.'),
    ]
    for fname, loc, desc in entries:
        already = False
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=1).value
            if val and fname in str(val):
                already = True
                break
        if not already:
            r = ws.max_row + 1
            ws.cell(row=r, column=1, value=fname)
            ws.cell(row=r, column=2, value=loc)
            ws.cell(row=r, column=3, value=desc)
            ws.cell(row=r, column=4, value=today)
    print('Documentacion entries added')

wb.save(EXCEL_PATH)
print(f'\nSaved: {EXCEL_PATH}')
print('Done.')
