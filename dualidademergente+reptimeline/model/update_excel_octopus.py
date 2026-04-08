"""Update Excel with cross-model 'octopus crystallization' confirmation."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
from datetime import date
import openpyxl
from openpyxl.styles import PatternFill, Font

EXCEL_PATH = r'C:\Github\7_caminos\seven-bridges\analisis_cruzado\analisis_completo_danza_cosmica.xlsx'

GREEN = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
YELLOW = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
BOLD = Font(bold=True)
today = date.today().isoformat()

wb = openpyxl.load_workbook(EXCEL_PATH)

# ──────────────────────────────────────────────────────────────
#  1. Update Status formalizacion — Q5 v9 result (add octopus note)
# ──────────────────────────────────────────────────────────────
ws = wb['Status formalizacion']

probes_start = None
for row in range(1, ws.max_row + 1):
    val = ws.cell(row=row, column=1).value
    if val and 'SECCION D' in str(val):
        probes_start = row
        break

if probes_start:
    header_row = probes_start + 1
    v9_col = None
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val and 'v9' in str(val).lower():
            v9_col = col
            break

    if v9_col:
        for row in range(header_row + 1, ws.max_row + 1):
            val = ws.cell(row=row, column=1).value
            if val and 'Q5' in str(val):
                old = str(ws.cell(row=row, column=v9_col).value or '')
                ws.cell(row=row, column=v9_col,
                    value=old + ' | CRISTALIZACION REPLICADA: v8 (inerte 5K-25K, explosion 35K, '
                          'churn 1.0->0.489) y v9 (explosion 10K, churn 1.0->0.458). '
                          'Patron "pulpo": organizacion interna -> cristalizacion gradual. '
                          'Misma dinamica en ambos modelos.')
                ws.cell(row=row, column=v9_col).fill = GREEN
                print(f'Updated Q5 at row {row}')
                break

    # Also update C1 conjecture with cross-model confirmation
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val and str(val).strip() == 'C1':
            for col in range(6, ws.max_column + 1):
                cell_val = ws.cell(row=row, column=col).value
                if cell_val and ('PROBADA' in str(cell_val) or 'v9' in str(cell_val)):
                    current = str(cell_val)
                    if 'pulpo' not in current:
                        ws.cell(row=row, column=col,
                            value=current + f' | Cristalizacion cross-model ({today}): '
                                  f'patron "pulpo" (reorganizacion->estabilizacion) '
                                  f'confirmado en v8 y v9.')
                    break
            break

    print('Updated C1 conjecture')

# ──────────────────────────────────────────────────────────────
#  2. Update Debilidades #5 (N=1) — crystallization replication
# ──────────────────────────────────────────────────────────────
ws = wb['Debilidades']

status_col = None
for col in range(1, ws.max_column + 1):
    for check_row in [1, 2]:
        val = ws.cell(row=check_row, column=col).value
        if val and 'Status' in str(val):
            status_col = col
            break
    if status_col:
        break

if status_col:
    for row in range(2, ws.max_row + 1):
        num = ws.cell(row=row, column=1).value
        if num == 5:
            old = str(ws.cell(row=row, column=status_col).value or '')
            if 'cristalizacion' not in old.lower():
                ws.cell(row=row, column=status_col,
                    value=old + f' | Cristalizacion ({today}): Patron "pulpo" '
                          f'(explosion churn=1.0 -> descenso gradual) replicado en ambos modelos. '
                          f'v8: transicion step 35K, v9: step 10K. Dinamica identica.')
            break
    print('Updated Debilidades #5')

# ──────────────────────────────────────────────────────────────
#  3. Update Roadmap — note about cross-model crystallization
# ──────────────────────────────────────────────────────────────
ws = wb['Roadmap']

status_col = None
for col in range(1, ws.max_column + 1):
    for check_row in [1, 2, 3]:
        val = ws.cell(row=check_row, column=col).value
        if val and 'Status' in str(val):
            status_col = col
            break
    if status_col:
        break

if status_col:
    next_row = ws.max_row + 1
    items = [
        (36, 'Confirmacion cross-model: patron "pulpo" de cristalizacion',
         'Churn reptimeline: ambos modelos muestran explosion (churn=1.0) '
         'seguida de cristalizacion gradual. v8: inerte 5K-25K, explosion 35K, '
         'descenso a 0.489. v9: explosion 10K, descenso a 0.458. '
         'Patron identico pese a diferente arch/dataset. '
         'Documentado en paper como "Inverted layer emergence" (sec 4.2 linea 385).',
         'HECHO', '#33, #34', 'P3', '—',
         'results/v8_curves.csv, results/v9_curves.csv',
         f'COMPLETADO {today}'),
    ]

    for item in items:
        for col_idx, val in enumerate(item, 1):
            cell = ws.cell(row=next_row, column=col_idx, value=val)
            if 'COMPLETADO' in str(val):
                cell.fill = GREEN
        next_row += 1

    print('Added Roadmap #36')

# ──────────────────────────────────────────────────────────────
#  4. Update Papers P3 — add crystallization replication
# ──────────────────────────────────────────────────────────────
ws = wb['Papers']
for row in range(2, ws.max_row + 1):
    val = ws.cell(row=row, column=1).value
    if val and str(val).strip() == 'P3':
        current = str(ws.cell(row=row, column=5).value or '')
        if 'pulpo' not in current:
            ws.cell(row=row, column=5,
                value=current + f' | Patron "pulpo" ({today}): cristalizacion '
                      f'(churn explosion->descenso) replicada cross-model. '
                      f'Inverted layer emergence confirmado en v9.')
        break

print('Updated Papers P3')

# ──────────────────────────────────────────────────────────────
wb.save(EXCEL_PATH)
print(f'\nSaved: {EXCEL_PATH}')
print('Done.')
