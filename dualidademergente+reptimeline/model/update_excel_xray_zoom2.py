"""Update Excel with V9_xray zoom2 results — cascade discovery."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
from datetime import date
import openpyxl
from openpyxl.styles import PatternFill, Font

EXCEL_PATH = r'C:\Github\7_caminos\seven-bridges\analisis_cruzado\analisis_completo_danza_cosmica.xlsx'
GREEN = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
YELLOW = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
today = date.today().isoformat()

wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb['Roadmap']

status_col = None
for col in range(1, ws.max_column + 1):
    for check_row in [1, 2, 3]:
        val = ws.cell(row=check_row, column=col).value
        if val and 'Status' in str(val):
            status_col = col
            break
    if status_col:
        break

if status_col:
    for row in range(3, ws.max_row + 1):
        num = ws.cell(row=row, column=1).value
        if num == 40:
            ws.cell(row=row, column=status_col,
                value=f'COMPLETADO {today}. HALLAZGO: la "explosion" (churn 0->1.0) era artefacto de '
                      f'baja resolucion. A 5-step res, churn sube gradualmente (~0.17). '
                      f'Los bits cambian en CASCADA (oleadas de 1-8 bits en ~45 steps). '
                      f'6 bits semilla en step 5060: [12,41,23,47,31,35]. '
                      f'31 bits total cambian en oleadas hasta step 5100. '
                      f'No es transicion de fase instantanea — es onda de activacion.')
            ws.cell(row=row, column=status_col).fill = GREEN
        elif num == 39:
            ws.cell(row=row, column=status_col,
                value=f'REFORMULAR ({today}): V9_xray zoom2 revela que la emergencia '
                      f'NO es simultanea ni secuencial (L1->L6). Es una CASCADA: '
                      f'6 bits semilla disparan una onda que recluta bits adicionales en oleadas. '
                      f'Pendiente: mapear bits semilla a capas/primitivos de la ontologia.')
            ws.cell(row=row, column=status_col).fill = YELLOW

    print('Updated Roadmap #39, #40')

# Update Q5 with cascade finding
ws = wb['Status formalizacion']
probes_start = None
for row in range(1, ws.max_row + 1):
    val = ws.cell(row=row, column=1).value
    if val and 'SECCION D' in str(val):
        probes_start = row
        break

if probes_start:
    header_row = probes_start + 1
    v9_col = None
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val and 'v9' in str(val).lower():
            v9_col = col
            break

    if v9_col:
        for row in range(header_row + 1, ws.max_row + 1):
            val = ws.cell(row=row, column=1).value
            if val and 'Q5' in str(val):
                old = str(ws.cell(row=row, column=v9_col).value or '')
                if 'cascada' not in old.lower():
                    ws.cell(row=row, column=v9_col,
                        value=old + f' | XRAY ZOOM2 ({today}): La "explosion" es una CASCADA, '
                              f'no transicion instantanea. 6 bits semilla (step 5060) disparan '
                              f'oleadas de 1-8 bits en ~45 steps. Churn real ~0.17, no 1.0.')
                break

        # Update Q1 with new interpretation
        for row in range(header_row + 1, ws.max_row + 1):
            val = ws.cell(row=row, column=1).value
            if val and 'Q1' in str(val):
                old = str(ws.cell(row=row, column=v9_col).value or '')
                if 'cascada' not in old.lower():
                    ws.cell(row=row, column=v9_col,
                        value=old + f' | XRAY ({today}): Emergencia NO es simultanea ni L1->L6. '
                              f'Es cascada: bits semilla [12,41,23,47,31,35] -> oleadas. '
                              f'Reformular Q1: ¿que bits/capas inician la cascada?')
                break

    print('Updated Q5 and Q1 with cascade finding')

wb.save(EXCEL_PATH)
print(f'\nSaved: {EXCEL_PATH}')
print('Done.')
