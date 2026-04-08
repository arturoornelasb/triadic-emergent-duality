"""Update Excel with cascade-to-layer mapping from V9_xray_zoom2."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
from datetime import date
import openpyxl
from openpyxl.styles import PatternFill, Font

EXCEL_PATH = r'C:\Github\7_caminos\seven-bridges\analisis_cruzado\analisis_completo_danza_cosmica.xlsx'
GREEN = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
BLUE = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
today = date.today().isoformat()

wb = openpyxl.load_workbook(EXCEL_PATH)

# --- 1. Update Roadmap: add #41 for cascade-to-layer mapping ---
ws = wb['Roadmap']
status_col = None
for col in range(1, ws.max_column + 1):
    for check_row in [1, 2, 3]:
        val = ws.cell(row=check_row, column=col).value
        if val and 'Status' in str(val):
            status_col = col
            break
    if status_col:
        break

if status_col:
    # Find last used row
    last_row = ws.max_row
    new_row = last_row + 1
    # Check if #41 already exists
    exists = False
    for row in range(3, last_row + 1):
        num = ws.cell(row=row, column=1).value
        if num == 41:
            exists = True
            ws.cell(row=row, column=status_col,
                value=f'COMPLETADO {today}. Cascade bits mapeados a capas ontologicas. '
                      f'Orden de activacion: L3-L5 (centro) -> L2 -> L6 -> L1 (extremos). '
                      f'6 bits semilla (step 5060): flujo_temporal(L3), colectivo(L5), caos(L3), '
                      f'algunos(L4), control(L4), dolor(L5). '
                      f'Ultimos 3 bits (step 5100): vacio(L1), uno(L1), eterno_obs(L6). '
                      f'5 bits estables (scaffold): tipo_de(L4), proporcion(L3), cooperacion(L4), '
                      f'atencion(L5), intencion(L5). '
                      f'La cascada va de capas intermedias a extremos ontologicos.')
            ws.cell(row=row, column=status_col).fill = GREEN
            break

    if not exists:
        ws.cell(row=new_row, column=1, value=41)
        # Find description column (usually col 2)
        ws.cell(row=new_row, column=2,
            value='Mapear bits de la cascada a capas/primitivos de la ontologia')
        ws.cell(row=new_row, column=status_col,
            value=f'COMPLETADO {today}. Cascade bits mapeados a capas ontologicas. '
                  f'Orden de activacion: L3-L5 (centro) -> L2 -> L6 -> L1 (extremos). '
                  f'6 bits semilla (step 5060): flujo_temporal(L3), colectivo(L5), caos(L3), '
                  f'algunos(L4), control(L4), dolor(L5). '
                  f'Ultimos 3 bits (step 5100): vacio(L1), uno(L1), eterno_obs(L6). '
                  f'5 bits estables (scaffold): tipo_de(L4), proporcion(L3), cooperacion(L4), '
                  f'atencion(L5), intencion(L5). '
                  f'La cascada va de capas intermedias a extremos ontologicos.')
        ws.cell(row=new_row, column=status_col).fill = GREEN
    print(f'Roadmap #41 {"updated" if exists else "added"}')

    # Update #39 with layer-order detail
    for row in range(3, ws.max_row + 1):
        num = ws.cell(row=row, column=1).value
        if num == 39:
            old = str(ws.cell(row=row, column=status_col).value or '')
            if 'L3-L5' not in old:
                ws.cell(row=row, column=status_col,
                    value=old + f' | MAPEO ({today}): Orden confirmado: L3-L5 (centro) -> L2 -> L6 -> L1. '
                          f'No es L1->L6 ni L6->L1. Es centro-afuera.')
            break
    print('Roadmap #39 updated with layer order')

# --- 2. Update Status formalizacion Q1 with definitive cascade order ---
ws = wb['Status formalizacion']
probes_start = None
for row in range(1, ws.max_row + 1):
    val = ws.cell(row=row, column=1).value
    if val and 'SECCION D' in str(val):
        probes_start = row
        break

if probes_start:
    header_row = probes_start + 1
    v9_col = None
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val and 'v9' in str(val).lower():
            v9_col = col
            break

    if v9_col:
        for row in range(header_row + 1, ws.max_row + 1):
            val = ws.cell(row=row, column=1).value
            if val and 'Q1' in str(val):
                old = str(ws.cell(row=row, column=v9_col).value or '')
                if 'L3-L5' not in old:
                    ws.cell(row=row, column=v9_col,
                        value=old + f' | MAPEO ({today}): Cascade order = L3-L5 -> L2 -> L6 -> L1. '
                              f'Centro ontologico (tiempo,moral,cuerpo) se activa primero; '
                              f'extremos (existencia,observador) al final. '
                              f'Reformular: emergencia es centro-afuera, no bottom-up ni top-down.')
                break
        print('Q1 updated with definitive cascade order')

# --- 3. Add documentation entry for map_cascade_bits.py ---
ws = wb['Documentacion'] if 'Documentacion' in wb.sheetnames else None
if ws:
    last_row = ws.max_row + 1
    # Check if already documented
    already = False
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val and 'map_cascade_bits' in str(val):
            already = True
            break
    if not already:
        ws.cell(row=last_row, column=1, value='map_cascade_bits.py')
        ws.cell(row=last_row, column=2, value='model/')
        ws.cell(row=last_row, column=3,
            value=f'Mapea bits de cascada (V9_xray_zoom2) a primitivos y capas ontologicas. '
                  f'Muestra distribucion de capas por oleada. '
                  f'Hallazgo: cascada L3-L5 -> L2 -> L6 -> L1.')
        ws.cell(row=last_row, column=4, value=today)
        print(f'Documentacion entry added for map_cascade_bits.py')
    else:
        print('map_cascade_bits.py already documented')
else:
    print('No Documentacion sheet found, skipping')

wb.save(EXCEL_PATH)
print(f'\nSaved: {EXCEL_PATH}')
print('Done.')
