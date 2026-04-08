"""Update Q8 result in analisis_completo_danza_cosmica.xlsx."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
from datetime import date
import openpyxl
from openpyxl.styles import PatternFill

EXCEL_PATH = r'C:\Github\7_caminos\seven-bridges\analisis_cruzado\analisis_completo_danza_cosmica.xlsx'
RED = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb['Status formalizacion']

# Find SECCION D header row
probes_start = None
for row in range(1, ws.max_row + 1):
    val = ws.cell(row=row, column=1).value
    if val and 'SECCION D' in str(val):
        probes_start = row
        break

if probes_start:
    header_row = probes_start + 1
    # Find v9 column
    v9_col = None
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val and 'v9' in str(val).lower():
            v9_col = col
            break

    if v9_col:
        # Find Q8 row and update
        for row in range(header_row + 1, ws.max_row + 1):
            val = ws.cell(row=row, column=1).value
            if val and 'Q8' in str(val):
                cell = ws.cell(row=row, column=v9_col,
                    value='NULL: Jaccard_duals=0.000, Jaccard_deps=0.052, SHD=147, NMI=0.000. '
                          'Cantidades similares (70vs72 active, 10vs9 duals, 81vs82 deps) '
                          'pero patrones especificos divergen entre GPT-Neo 125M y GPT-2 Medium.')
                cell.fill = RED
                print(f'Updated Q8 at row {row}, col {v9_col}')
                break
        else:
            print('Q8 row not found')
    else:
        print('v9 column not found')
else:
    print('SECCION D not found')

wb.save(EXCEL_PATH)
print(f'Saved: {EXCEL_PATH}')
