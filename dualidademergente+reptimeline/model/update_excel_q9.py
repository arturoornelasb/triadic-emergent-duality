"""Update Excel with Q9 structural conservation results + add Q9 row."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
from datetime import date
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

EXCEL_PATH = r'C:\Github\7_caminos\seven-bridges\analisis_cruzado\analisis_completo_danza_cosmica.xlsx'

GREEN = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
YELLOW = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
BLUE = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
BOLD = Font(bold=True)
today = date.today().isoformat()

wb = openpyxl.load_workbook(EXCEL_PATH)

# ──────────────────────────────────────────────────────────────
#  1. Add Q9 row to Status formalizacion — SECCION D
# ──────────────────────────────────────────────────────────────
ws = wb['Status formalizacion']

probes_start = None
for row in range(1, ws.max_row + 1):
    val = ws.cell(row=row, column=1).value
    if val and 'SECCION D' in str(val):
        probes_start = row
        break

if probes_start:
    header_row = probes_start + 1

    # Find v9 column
    v9_col = None
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val and 'v9' in str(val).lower():
            v9_col = col
            break

    # Find last Q row in section D
    last_q_row = header_row
    for row in range(header_row + 1, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val and str(val).strip().startswith('Q'):
            last_q_row = row
        elif val and ('SECCION' in str(val) or str(val).strip() == ''):
            break

    # Add Q9 row after last Q row
    q9_row = last_q_row + 1
    ws.cell(row=q9_row, column=1,
            value='Q9: Conservacion estructural cross-model')
    ws.cell(row=q9_row, column=1).font = BOLD

    # Fill description columns (same as other Q rows)
    # Col 2: description
    ws.cell(row=q9_row, column=2,
            value='Compara propiedades distribucionales (no bit-identity) entre modelos: '
                  'active ratio, duals, deps, densidad, distribucion S/C/E/A. '
                  'Complementa Q8 (que mide identidad especifica).')

    # Col 3: theoretical prediction
    ws.cell(row=q9_row, column=3,
            value='Si la estructura triadica refleja la ontologia (72 primitivos), '
                  'las propiedades distribucionales deberian conservarse entre modelos '
                  'aunque los bits especificos difieran.')

    if v9_col:
        cell = ws.cell(row=q9_row, column=v9_col,
            value='MIXED: Topologia base conservada (active 97%, duals 90%, deps 99%, '
                  'densidad 96%). Triadic deps difieren por escala (838 vs 2462, 3x). '
                  'S/C/E/A chi2 p<0.001 (diverge). Modelo mayor = mas sinergia (42% vs 34%).')
        cell.fill = YELLOW

    print(f'Added Q9 row at {q9_row}')

# ──────────────────────────────────────────────────────────────
#  2. Update Roadmap — add Q9 item
# ──────────────────────────────────────────────────────────────
ws = wb['Roadmap']

status_col = None
for col in range(1, ws.max_column + 1):
    for check_row in [1, 2, 3]:
        val = ws.cell(row=check_row, column=col).value
        if val and 'Status' in str(val):
            status_col = col
            break
    if status_col:
        break

if status_col:
    next_row = ws.max_row + 1
    items = [
        (35, 'Q9: Conservacion estructural cross-model',
         'Nuevo probe: compara propiedades distribucionales (no bit-identity) '
         'entre v8 (GPT-2 Medium/TinyStories) y v9 (GPT-Neo 125M/OpenWebText). '
         'Topologia base conservada >90%. S/C/E/A difiere por escala (3x triadic deps en v8). '
         'Chi2 p<0.001.',
         'HECHO', '#33, #34', 'P3, P9', '—',
         'results/q9_structural_conservation.json', f'COMPLETADO {today}'),
    ]

    for item in items:
        for col_idx, val in enumerate(item, 1):
            cell = ws.cell(row=next_row, column=col_idx, value=val)
            if 'COMPLETADO' in str(val):
                cell.fill = GREEN
        next_row += 1

    print(f'Added Roadmap #35')

# ──────────────────────────────────────────────────────────────
#  3. Update Debilidades #5 (N=1) — Q9 adds nuance
# ──────────────────────────────────────────────────────────────
ws = wb['Debilidades']

status_col = None
for col in range(1, ws.max_column + 1):
    for check_row in [1, 2]:
        val = ws.cell(row=check_row, column=col).value
        if val and 'Status' in str(val):
            status_col = col
            break
    if status_col:
        break

if status_col:
    for row in range(2, ws.max_row + 1):
        num = ws.cell(row=row, column=1).value
        if num == 5:
            old = str(ws.cell(row=row, column=status_col).value or '')
            if 'Q9' not in old:
                ws.cell(row=row, column=status_col,
                    value=old + f' | Q9 ({today}): Topologia base conservada cross-model '
                          f'(active 97%, duals 90%, deps 99%). Estructura fundamental '
                          f'la dictan los 72 primitivos, no el backbone.')
            break

    print('Updated Debilidades #5 with Q9')

# ──────────────────────────────────────────────────────────────
#  Save
# ──────────────────────────────────────────────────────────────
wb.save(EXCEL_PATH)
print(f'\nSaved: {EXCEL_PATH}')
print('Done.')
