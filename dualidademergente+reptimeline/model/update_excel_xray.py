"""Update Excel with V9_xray experiment documentation."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
from datetime import date
import openpyxl
from openpyxl.styles import PatternFill, Font

EXCEL_PATH = r'C:\Github\7_caminos\seven-bridges\analisis_cruzado\analisis_completo_danza_cosmica.xlsx'

BLUE = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
YELLOW = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
GREEN = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
BOLD = Font(bold=True)
today = date.today().isoformat()

wb = openpyxl.load_workbook(EXCEL_PATH)

# ──────────────────────────────────────────────────────────────
#  1. Roadmap — add V9_xray items
# ──────────────────────────────────────────────────────────────
ws = wb['Roadmap']

status_col = None
for col in range(1, ws.max_column + 1):
    for check_row in [1, 2, 3]:
        val = ws.cell(row=check_row, column=col).value
        if val and 'Status' in str(val):
            status_col = col
            break
    if status_col:
        break

if status_col:
    next_row = ws.max_row + 1
    items = [
        (37, 'V9_xray Paso 1: radiografia transicion de fase (1K res)',
         'Reentrenar v9 desde step 5K hasta 15K guardando checkpoint cada 1000 steps. '
         'Objetivo: determinar si la explosion (churn 0->1.0) es instantanea o tiene '
         'micro-secuencia. Enfoque telescopico: paso 1 (1K res) -> paso 2 (100 step res). '
         'Resume desde v9_neo_openwebtext/step_5000.pt.',
         'EN PROGRESO', '#33', 'P3', '—',
         'checkpoints/v9_xray/, runs/v9_xray/plots/, results/v9_xray_*',
         f'INICIADO {today}'),
        (38, 'V9_xray Paso 2: zoom alta resolucion (100 step res)',
         'Reentrenar solo el rango exacto de la explosion (identificado en paso 1) '
         'con checkpoint cada 100 steps. Ejemplo: si explosion entre step 7K y 8K, '
         'reentrenar ese rango con --save-every 100. ~10 checkpoints adicionales.',
         'PENDIENTE', '#37', 'P3', '—',
         'checkpoints/v9_xray_zoom/', f'PENDIENTE'),
        (39, 'Reformular Q1 en paper: emergencia simultanea vs secuencial',
         'Q1 preguntaba si capas emergen en orden L1->L6. Datos muestran emergencia '
         'simultanea (transicion de fase). Reformular: la estructura emerge como un todo '
         'coherente (criticidad), no pieza por pieza. V9_xray determinara la granularidad '
         'real de "simultaneo".',
         'PENDIENTE', '#37, #38', 'P3', '—',
         'paper/triadic_duality.tex (linea 385)', 'PENDIENTE'),
    ]

    for item in items:
        for col_idx, val in enumerate(item, 1):
            cell = ws.cell(row=next_row, column=col_idx, value=val)
            if 'COMPLETADO' in str(val):
                cell.fill = GREEN
            elif 'EN PROGRESO' in str(val) or 'INICIADO' in str(val):
                cell.fill = YELLOW
            elif 'PENDIENTE' in str(val):
                cell.fill = BLUE
        next_row += 1

    print(f'Added Roadmap #37-#39')

# ──────────────────────────────────────────────────────────────
#  2. Documentacion — update with version log reference
# ──────────────────────────────────────────────────────────────
ws = wb['Documentacion']

# Find last row with content
last_row = ws.max_row + 2

doc_entries = [
    ('Version Log (todo el programa)',
     'V7_PLAN.md en dualidademergente+reptimeline/. '
     'Documenta V2-V9 + V9_xray: cambios, hiperparametros, resultados, '
     'diagnosticos reptimeline, archivos generados. '
     'Es el registro maestro de reproducibilidad.'),
    ('Archivos de actualizacion Excel',
     'model/update_excel_v9.py — Resultados v9 Q1-Q7 + FEP + debilidades. '
     'model/update_excel_q8.py — Q8 cross-arch identity. '
     'model/update_excel_q9.py — Q9 conservacion estructural. '
     'model/update_excel_octopus.py — Cristalizacion cross-model. '
     'model/update_excel_xray.py — V9_xray y documentacion. '
     'Todos son reproducibles y NO modifican datos, solo agregan.'),
    ('Indice de scripts por version',
     'V8: train.py + save_v8_timeline.py + run_v8.bat. '
     'V9: train_v9_neo.bat + save_v9_timeline.py + fit_wave_fep_v9.py + eval_v9_full.bat. '
     'V9_xray: train_v9_xray.bat + save_v9_xray_timeline.py. '
     'Probes: neural/q1-q9_probe_*.py + run_all_probes_v9.bat. '
     'Todos en dualidademergente+reptimeline/.'),
    ('Estructura de datos para papers',
     'Checkpoints: model/checkpoints/{v8_deep, v9_neo_openwebtext, v9_xray}/. '
     'Resultados: neural/results/{q1-q9_*.json, v8_*.json, v9_*.json, v9_xray_*.json}. '
     'Graficas: runs/{v8_deep, v9_neo, v9_xray}/plots/. '
     'Paper: paper/triadic_duality.tex. '
     'Excel maestro: analisis_cruzado/analisis_completo_danza_cosmica.xlsx.'),
]

for title, desc in doc_entries:
    ws.cell(row=last_row, column=1, value=title)
    ws.cell(row=last_row, column=1).font = BOLD
    ws.cell(row=last_row, column=2, value=desc)
    last_row += 1

ws.cell(row=last_row, column=1, value=f'Ultima actualizacion documentacion: {today}')

print('Updated Documentacion sheet')

# ──────────────────────────────────────────────────────────────
#  Save
# ──────────────────────────────────────────────────────────────
wb.save(EXCEL_PATH)
print(f'\nSaved: {EXCEL_PATH}')
print('Done.')
