"""Update Excel with V8_xray paso 1 results."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
from datetime import date
import openpyxl
from openpyxl.styles import PatternFill

EXCEL_PATH = r'C:\Github\7_caminos\seven-bridges\analisis_cruzado\analisis_completo_danza_cosmica.xlsx'
GREEN = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
YELLOW = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
today = date.today().isoformat()

wb = openpyxl.load_workbook(EXCEL_PATH)

# --- 1. Update Roadmap ---
ws = wb['Roadmap']
status_col = None
for col in range(1, ws.max_column + 1):
    for check_row in [1, 2, 3]:
        val = ws.cell(row=check_row, column=col).value
        if val and 'Status' in str(val):
            status_col = col
            break
    if status_col:
        break

if status_col:
    # Add #42: V8_xray paso 1
    last_row = ws.max_row
    exists = False
    for row in range(3, last_row + 1):
        num = ws.cell(row=row, column=1).value
        if num == 42:
            exists = True
            target_row = row
            break
    if not exists:
        target_row = last_row + 1
        ws.cell(row=target_row, column=1, value=42)
        ws.cell(row=target_row, column=2,
            value='V8_xray paso 1: radiografia transicion de fase en v8 (GPT-2 Medium)')

    ws.cell(row=target_row, column=status_col,
        value=f'COMPLETADO {today}. Explosion localizada entre step 31K-32K '
              f'(NO 35K como se reportaba). Churn 0->0.81, nunca 1.0. '
              f'Patron identico a v9: explosion seguida de cristalizacion gradual. '
              f'15 checkpoints en v8_xray/. Zoom (paso 2) en progreso.')
    ws.cell(row=target_row, column=status_col).fill = GREEN
    print(f'Roadmap #42 {"updated" if exists else "added"}')

    # Add #43: V8_xray zoom (in progress)
    exists43 = False
    for row in range(3, ws.max_row + 1):
        num = ws.cell(row=row, column=1).value
        if num == 43:
            exists43 = True
            target_row43 = row
            break
    if not exists43:
        target_row43 = ws.max_row + 1
        ws.cell(row=target_row43, column=1, value=43)
        ws.cell(row=target_row43, column=2,
            value='V8_xray paso 2: zoom 50-step en explosion (30K-32K)')
    ws.cell(row=target_row43, column=status_col,
        value=f'EN PROGRESO {today}. Resume desde step_30000, '
              f'save-every 50 hasta 32K (40 checkpoints). '
              f'Objetivo: localizar rango exacto para zoom2.')
    ws.cell(row=target_row43, column=status_col).fill = YELLOW
    print(f'Roadmap #43 {"updated" if exists43 else "added"}')

# --- 2. Update Documentacion ---
ws = wb['Documentacion'] if 'Documentacion' in wb.sheetnames else None
if ws:
    entries = [
        ('train_v8_xray.bat', 'model/',
         f'Paso 1 V8_xray: 30K->45K cada 1000 steps. GPT-2 Medium, resume from v8_deep/step_30000.pt'),
        ('save_v8_xray_timeline.py', 'model/',
         f'Extraccion reptimeline para V8_xray paso 1. Genera v8_xray_curves.csv y plots.'),
        ('train_v8_xray_zoom.bat', 'model/',
         f'Paso 2 V8_xray: 30K->32K cada 50 steps. Zoom en la explosion.'),
        ('save_v8_xray_zoom_timeline.py', 'model/',
         f'Extraccion reptimeline para V8_xray paso 2.'),
    ]
    for fname, loc, desc in entries:
        already = False
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=1).value
            if val and fname in str(val):
                already = True
                break
        if not already:
            r = ws.max_row + 1
            ws.cell(row=r, column=1, value=fname)
            ws.cell(row=r, column=2, value=loc)
            ws.cell(row=r, column=3, value=desc)
            ws.cell(row=r, column=4, value=today)
    print('Documentacion entries added')

wb.save(EXCEL_PATH)
print(f'\nSaved: {EXCEL_PATH}')
print('Done.')
