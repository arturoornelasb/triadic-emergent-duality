"""Update analisis_completo_danza_cosmica.xlsx with v9 results."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import os
import json
from datetime import date
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

EXCEL_PATH = r'C:\Github\7_caminos\seven-bridges\analisis_cruzado\analisis_completo_danza_cosmica.xlsx'
RESULTS_DIR = r'C:\Github\dualidad_emergente\dualidademergente+reptimeline\neural\results'

# Colors
GREEN = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
YELLOW = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
BLUE = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
BOLD = Font(bold=True)

wb = openpyxl.load_workbook(EXCEL_PATH)
today = date.today().isoformat()

# ─── Load v9 results ───
fep_fit = {}
fep_path = os.path.join(RESULTS_DIR, 'v9_fep_wave_fit.json')
if os.path.exists(fep_path):
    with open(fep_path) as f:
        fep_fit = json.load(f)

# ──────────────────────────────────────────────────────────────
#  1. Update Documentacion sheet (last updated date)
# ──────────────────────────────────────────────────────────────
ws = wb['Documentacion']
ws['A3'] = f'Ultima actualizacion: {today}'
print('Updated: Documentacion (date)')

# ──────────────────────────────────────────────────────────────
#  2. Update Status formalizacion — Neural Probes section
# ──────────────────────────────────────────────────────────────
ws = wb['Status formalizacion']

# Find the Neural Probes section (Section D)
probes_start = None
for row in range(1, ws.max_row + 1):
    val = ws.cell(row=row, column=1).value
    if val and 'SECCION D' in str(val):
        probes_start = row
        break

if probes_start:
    # Find header row
    header_row = probes_start + 1
    # Add v9 columns after existing ones
    # Find existing columns
    cols = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val:
            cols[str(val).strip()] = col

    # Add v9 Result column after v8 Result (or at end)
    v9_col = ws.max_column + 1
    for col_name, col_idx in cols.items():
        if 'v8' in col_name.lower():
            v9_col = col_idx + 1
            break

    # Insert v9 results column header
    ws.cell(row=header_row, column=v9_col, value='v9 Result (GPT-Neo 125M/OWT)')
    ws.cell(row=header_row, column=v9_col).font = BOLD
    ws.cell(row=header_row, column=v9_col).fill = BLUE

    # Map probe results
    v9_results = {
        'Q1': 'NULL (rho=-0.072, p=0.55)',
        'Q2': 'NULL (1/14 anti-corr)',
        'Q3': 'NULL (p=0.001 vs random, F1=0.152)',
        'Q4': 'MIXED (66% complementary, 0.6% emergent)',
        'Q5': 'POSITIVE (p=0.0182) — REPLICADO',
        'Q6': 'MIXED (77.8% agreement)',
        'Q7': 'NULL (40/72 sig, no layer effect)',
        'Q8': 'N/A (necesita 2 archs, probe rediseñar para v8 vs v9)',
    }

    # Find probe rows and fill v9 results
    for row in range(header_row + 1, ws.max_row + 1):
        probe_val = ws.cell(row=row, column=1).value
        if probe_val:
            probe_str = str(probe_val).strip()
            for qname, result in v9_results.items():
                if probe_str.startswith(qname) or qname in probe_str:
                    cell = ws.cell(row=row, column=v9_col, value=result)
                    if 'POSITIVE' in result or 'REPLICADO' in result:
                        cell.fill = GREEN
                    elif 'MIXED' in result:
                        cell.fill = YELLOW
                    break

    print(f'Updated: Status formalizacion (v9 results in col {v9_col})')

# ──────────────────────────────────────────────────────────────
#  3. Update Debilidades sheet
# ──────────────────────────────────────────────────────────────
ws = wb['Debilidades']

# Find Status column
status_col = None
for col in range(1, ws.max_column + 1):
    val = ws.cell(row=1, column=col).value
    if val and 'Status' in str(val):
        status_col = col
        break

if not status_col:
    # Try row 2
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=2, column=col).value
        if val and 'Status' in str(val):
            status_col = col
            break

if status_col:
    for row in range(2, ws.max_row + 1):
        num = ws.cell(row=row, column=1).value
        if num == 5:  # N=1
            old = str(ws.cell(row=row, column=status_col).value or '')
            ws.cell(row=row, column=status_col,
                    value=f'PARCIALMENTE RESUELTA ({today}): v9 (GPT-Neo 125M, OpenWebText, 50K steps) entrenado y evaluado. '
                          f'Q5 REPLICADO (p=0.018). FEP wave fit genuino (f=1.45, dAIC=44.5). '
                          f'Falta: tercer modelo para replicación completa.')
            ws.cell(row=row, column=status_col).fill = YELLOW
        elif num == 6:  # Q1/Q2 NULL
            ws.cell(row=row, column=status_col,
                    value=f'CONFIRMADO NULL en v9 ({today}): Q1 (p=0.55) y Q2 (1/14) siguen NULL en GPT-Neo 125M. '
                          f'Escala 3x (40M→125M) insuficiente. Necesita 400M+ o revisión teórica.')
        elif num == 7:  # FEP ratio artefacto
            old = str(ws.cell(row=row, column=status_col).value or '')
            ws.cell(row=row, column=status_col,
                    value=f'RESUELTA en v9 ({today}): Wave fit R²=0.734, dAIC=44.5 (wave >> exponencial). '
                          f'f=1.45 CI[0.49,2.49], régimen underdamped (d/ω=0.703). '
                          f'v9 muestra oscilación genuina, no artefacto. Bootstrap 1000/1000 exitoso.')
            ws.cell(row=row, column=status_col).fill = GREEN
    print(f'Updated: Debilidades (#5, #6, #7)')

# ──────────────────────────────────────────────────────────────
#  4. Update Roadmap sheet
# ──────────────────────────────────────────────────────────────
ws = wb['Roadmap']

# Find Status column
status_col = None
for col in range(1, ws.max_column + 1):
    for check_row in [1, 2, 3]:
        val = ws.cell(row=check_row, column=col).value
        if val and 'Status' in str(val):
            status_col = col
            break
    if status_col:
        break

if status_col:
    for row in range(3, ws.max_row + 1):
        num = ws.cell(row=row, column=1).value
        if num == 12:  # Fit onda a datos reptimeline
            ws.cell(row=row, column=status_col,
                    value=f'COMPLETADO {today}. Wave fit R²=0.734, underdamped (d/ω=0.703), f=1.45, dAIC=44.5. '
                          f'Oscilación genuina confirmada. fit_wave_fep_v9.py')
            ws.cell(row=row, column=status_col).fill = GREEN
        elif num == 14:  # Escalar TriadicGPT
            old = str(ws.cell(row=row, column=status_col).value or '')
            if 'COMPLETADO' not in old:
                ws.cell(row=row, column=status_col,
                        value=f'PARCIAL {today}: v9 (GPT-Neo 125M, OpenWebText) completado. '
                              f'bit_acc_test=83.2%, sub_test=93.1%. Q5 replicado. No es TriadicGPT escalado sino modelo alterno (N>1).')
                ws.cell(row=row, column=status_col).fill = YELLOW

    # Add new rows for v9 work
    next_row = ws.max_row + 2
    new_items = [
        (33, 'Entrenamiento v9 (GPT-Neo 125M)',
         'GPT-Neo 125M, OpenWebText 100K docs, 50K steps, 72 bits, deep head, IFSQ, freeze-base. '
         'bit_acc_test=83.2%, sub_test=93.1%, 30 dead bits, entropy=0.491.',
         'HECHO', '—', 'P2, P3, P5', '—',
         'checkpoints/v9_neo_openwebtext/', f'COMPLETADO {today}'),
        (34, 'Evaluación v9: reptimeline + probes Q1-Q8',
         '10 checkpoints analizados. 2 transiciones de fase (step 10K). '
         'Q5 POSITIVE (p=0.018) — replicación cross-model. Q1/Q2 NULL. '
         'Wave fit: R²=0.734, underdamped, f=1.45, dAIC=44.5 (genuino).',
         'HECHO', '#33', 'P3, P6, P9', '—',
         'results/v9_*, runs/v9_neo/plots/', f'COMPLETADO {today}'),
    ]

    for item in new_items:
        for col_idx, val in enumerate(item, 1):
            cell = ws.cell(row=next_row, column=col_idx, value=val)
            if 'COMPLETADO' in str(val):
                cell.fill = GREEN
        next_row += 1

    print(f'Updated: Roadmap (#12, #14, added #33-#34)')

# ──────────────────────────────────────────────────────────────
#  5. Update Conjectures (C1 replication check)
# ──────────────────────────────────────────────────────────────
ws = wb['Status formalizacion']
for row in range(1, ws.max_row + 1):
    val = ws.cell(row=row, column=1).value
    if val and str(val).strip() == 'C1':
        # Find the status/notes column
        for col in range(6, ws.max_column + 1):
            cell_val = ws.cell(row=row, column=col).value
            if cell_val and 'PROBADA' in str(cell_val):
                current = str(cell_val)
                ws.cell(row=row, column=col,
                        value=current + f' | v9 ({today}): Wave fit R²=0.734, underdamped, '
                              f'oscilación genuina (f=1.45). dAIC=44.5 vs exponencial.')
                break
        break

print('Updated: Conjetura C1 (v9 replication note)')

# ──────────────────────────────────────────────────────────────
#  6. Add v9 summary to Papers sheet (P3 notes)
# ──────────────────────────────────────────────────────────────
ws = wb['Papers']
for row in range(2, ws.max_row + 1):
    val = ws.cell(row=row, column=1).value
    if val and str(val).strip() == 'P3':
        # Find results column (col 5 = Resultados probados)
        current = str(ws.cell(row=row, column=5).value or '')
        ws.cell(row=row, column=5,
                value=current + f' | v9 REPLICACIÓN ({today}): GPT-Neo 125M/OpenWebText, '
                      f'2 transiciones de fase (step 10K). Q5 POSITIVE (p=0.018). '
                      f'Cristalización confirmada cross-model.')
        break

print('Updated: Papers P3 (v9 replication)')

# ──────────────────────────────────────────────────────────────
#  Save
# ──────────────────────────────────────────────────────────────
wb.save(EXCEL_PATH)
print(f'\nSaved: {EXCEL_PATH}')
print('Done.')
