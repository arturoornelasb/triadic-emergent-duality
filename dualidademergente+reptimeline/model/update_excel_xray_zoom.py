"""Update Excel with V9_xray zoom results and zoom2 planning."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
from datetime import date
import openpyxl
from openpyxl.styles import PatternFill, Font

EXCEL_PATH = r'C:\Github\7_caminos\seven-bridges\analisis_cruzado\analisis_completo_danza_cosmica.xlsx'
GREEN = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
YELLOW = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
BLUE = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
today = date.today().isoformat()

wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb['Roadmap']

# Update #37 status
status_col = None
for col in range(1, ws.max_column + 1):
    for check_row in [1, 2, 3]:
        val = ws.cell(row=check_row, column=col).value
        if val and 'Status' in str(val):
            status_col = col
            break
    if status_col:
        break

if status_col:
    for row in range(3, ws.max_row + 1):
        num = ws.cell(row=row, column=1).value
        if num == 37:
            ws.cell(row=row, column=status_col,
                value=f'COMPLETADO {today}. Explosion localizada: churn 0->1.0 entre '
                      f'step 5050-5100 (50 steps). Paso 2 (zoom 50-step res) completado: '
                      f'sub explota 1%->87% en 100 steps post-churn. '
                      f'Paso 3 (zoom2, 5-step res) en progreso.')
            ws.cell(row=row, column=status_col).fill = GREEN
        elif num == 38:
            ws.cell(row=row, column=status_col,
                value=f'COMPLETADO {today}. 20 checkpoints (5050-6000). '
                      f'Micro-secuencia: 1) churn explosion instantanea (5050-5100), '
                      f'2) sub sigmoide (5100-5200), 3) cristalizacion (5200-6000). '
                      f'Churn heatmap muestra bits con diferentes velocidades de estabilizacion.')
            ws.cell(row=row, column=status_col).fill = GREEN
            # Update description to match what was actually done
            ws.cell(row=row, column=3,
                value='Reentrenar rango 5K-6K con checkpoint cada 50 steps. '
                      '20 checkpoints. Resultado: churn 0->1.0 instantaneo (50 steps), '
                      'sub explosion sigmoide (100 steps), cristalizacion gradual.')

    # Add #40 for zoom2 (paso 3)
    next_row = ws.max_row + 1
    items = [
        (40, 'V9_xray Paso 3: zoom2 maxima resolucion (5 step res)',
         'Reentrenar step 5050 a 5100 con checkpoint cada 5 steps. '
         'Objetivo: determinar si churn 0->1.0 es verdaderamente instantaneo '
         'o tiene rampa interna. Resume desde v9_xray_zoom/step_5050.pt.',
         'EN PROGRESO', '#38', 'P3', '—',
         'checkpoints/v9_xray_zoom2/, runs/v9_xray_zoom2/plots/, results/v9_xray_zoom2_*',
         f'INICIADO {today}'),
    ]
    for item in items:
        for col_idx, val in enumerate(item, 1):
            cell = ws.cell(row=next_row, column=col_idx, value=val)
            if 'COMPLETADO' in str(val):
                cell.fill = GREEN
            elif 'EN PROGRESO' in str(val) or 'INICIADO' in str(val):
                cell.fill = YELLOW
        next_row += 1

    print('Updated Roadmap #37, #38, added #40')

wb.save(EXCEL_PATH)
print(f'Saved: {EXCEL_PATH}')
